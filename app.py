from flask import Flask, render_template, request, redirect, url_for, flash
from openpyxl import Workbook, load_workbook
import os
import pandas as pd
from statsmodels.tsa.arima.model import ARIMA
import matplotlib.pyplot as plt
from datetime import timedelta
from scipy.ndimage import gaussian_filter1d
import pickle

app = Flask(__name__)
app.secret_key = 'your_secret_key'

EXCEL_FILE = 'Datasets/User_Information/user_data.xlsx'

# Mapping of Item Name to Item Code
ITEM_CODE_MAPPING = {
    "Allgäuer Hof-Milch Butter mild gesäuert 250g": "101",
    "Bio Aubergine 1 Stück": "201",
    "Blumenkohl weiß 1 Stück": "301",
    "Broccoli 500g": "401",
    "Eisbergsalat 1 Stück": "501",
    "Galiamelone 1 Stück": "601",
    "Karotten 1kg": "701",
    "Kartoffeln vorwiegend festkochend 2,5kg": "801",
    "Mango vorgereift 1 Stück": "901",
    "Meggle Feine Butter 250g": "1001",
    "Orangen 2kg im Netz": "1101",
    "REWE Beste Wahl Feinschmecker Hähnchen 1200g": "1201",
    "REWE Bio Zucchini 500g": "1301",
    "Rewe Beste Wahl Eier aus Freilandhaltung 10 Stück": "1401",
    "Rispentomaten ca. 100g": "1501",
    "Spitzkohl ca. 1kg": "1601",
    "Tafeltrauben hell kernlos 500g": "1701",
    "Zitronen 500g im Netz": "1801",
    "Zwiebeln 2kg im Netz": "1901",
    "ja! Basmati Reis 1kg": "2001",
    "ja! H-Milch 3,5% 1": "2101",
    "ja! Sonnenblumenöl 1l": "2201"
}

# Mapping of Item Code to Item English Name
ITEM_CODE_TO_ENGLISH_NAME = {
    "101": "Butter Mildly Soured",
    "201": "Eggplant",
    "301": "White Cauliflower",
    "401": "Broccoli",
    "501": "Iceberg Lettuce",
    "601": "Galia Melon",
    "701": "Carrots",
    "801": "Potatoes",
    "901": "Mango",
    "1001": "Meggle Fine Butter",
    "1101": "Oranges",
    "1201": "Chicken",
    "1301": "Zucchini",
    "1401": "Eggs",
    "1501": "Tomatoes",
    "1601": "Cabbage",
    "1701": "Table Grapes",
    "1801": "Lemons",
    "1901": "Onions",
    "2001": "Basmati Rice",
    "2101": "Milk 3.5%",
    "2201": "Sunflower Oil"
}

def predict_and_save(item_name, models, data, steps=90, output_folder="Datasets/Output/Price_Prediction", csv_file="Datasets/Output/Price_Prediction/Item_lists.csv"):
    if item_name not in models:
        print(f"No model found for item: {item_name}")
        return

    # Create output directories
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    #item_folder = os.path.join(output_folder, item_name.replace(' ', '_'))
    item_folder = os.path.join(output_folder, 'Plots')
    
    if not os.path.exists(item_folder):
        os.makedirs(item_folder)

    model = models[item_name]
    item_data = data[data['Items'] == item_name]
    item_data['Date'] = pd.to_datetime(item_data['Date'])  # Ensure 'Date' is datetime
    daily_prices = item_data.set_index('Date')['price']

    # Forecast future values
    forecast = model.forecast(steps=steps)
    last_date = daily_prices.index[-1]
    future_dates = pd.date_range(start=last_date + timedelta(days=1), periods=steps)

    # Smooth historical prices
    smoothed_prices = gaussian_filter1d(daily_prices, sigma=2)

    # Smooth forecasted prices
    smoothed_forecast = gaussian_filter1d(forecast, sigma=2)

    # Combine historical and forecasted data
    combined_index = daily_prices.index.append(future_dates)
    combined_values = list(smoothed_prices) + list(smoothed_forecast)

    # Save forecast to CSV
    price_for_tomorrow = forecast.iloc[0]
    price_after_7_days = forecast.iloc[7] if len(forecast) > 7 else None
    price_after_1_month = forecast.iloc[30] if len(forecast) > 30 else None

    item_code = ITEM_CODE_MAPPING.get(item_name, "Unknown")
    item_english_name = ITEM_CODE_TO_ENGLISH_NAME.get(item_code, "Unknown")

    summary_data = {
        "Serial_Number": [len(pd.read_csv(csv_file)) + 1 if os.path.exists(csv_file) else 1],
        "Item_Code": [item_code],
        "Item_Name": [item_name],
        "Price_for_Tomorrow": [price_for_tomorrow],
        "Price_After_7_Days": [price_after_7_days],
        "Price_After_1_Month": [price_after_1_month]
    }

    summary_df = pd.DataFrame(summary_data)

    # Append to CSV if file exists, otherwise create
    if os.path.exists(csv_file):
        summary_df.to_csv(csv_file, mode='a', header=False, index=False)
    else:
        summary_df.to_csv(csv_file, index=False)

    
    print(f"Item's information added to {csv_file}")
    


    # Plot
    plt.figure(figsize=(12, 6))
    plt.plot(daily_prices.index, smoothed_prices, label="Actual Price", linewidth=2,linestyle='-', color="blue")
    plt.plot(future_dates, smoothed_forecast, label="Predicted Price (Next 90 Days)", linewidth=3,linestyle='-', color='orange')
    plt.title(f"Price Trend for {item_english_name}", fontsize=16)
    plt.ylabel("Price (€)")
    plt.legend()
    plt.grid(True)
    plt.tight_layout()

    # Save plot
    plot_file = os.path.join(item_folder, f"{item_code.replace(' ', '_')}.png")
    plt.savefig(plot_file)
    plt.close()
    print(f"Plot saved to {plot_file}")

# Initialize Excel file and ensure headers exist
def initialize_excel():
    if not os.path.exists(EXCEL_FILE):
        # Create a new Excel file with headers
        wb = Workbook()
        ws = wb.active
        ws.title = 'Users'
        ws.append(['Customer_ID', 'First_Name', 'Last_Name', 'Email', 'Country', 'City', 'Address', 'Password'])
        wb.save(EXCEL_FILE)
        print("Initialized 'user_data.xlsx' with headers.")


@app.route('/')
def login_page():
    return render_template('login.html')

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        # Collect form data
        first_name = request.form['first_name']
        last_name = request.form['last_name']
        email = request.form['email']
        country = request.form['country']
        city = request.form['city']
        address = request.form['address']
        password = request.form['password']

        # Load Excel file
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        # Ensure headers are intact before appending data
        if ws.cell(1, 1).value != 'Customer_ID':
            ws.insert_rows(1)
            ws.append(['Customer_ID', 'First_Name', 'Last_Name', 'Email', 'Country', 'City', 'Address', 'Password'])

        # Check for duplicate email
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[3] == email:  # Email is in the 4th column
                flash('Email already exists! Please log in.', 'error')
                return redirect(url_for('signup'))

        # Determine the next Customer_ID
        last_customer_id = 1000
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0] and isinstance(row[0], int):  # Ensure valid Customer_IDs
                last_customer_id = max(last_customer_id, int(row[0]))

        new_customer_id = last_customer_id + 1

        # Append new user data to Excel
        ws.append([new_customer_id, first_name, last_name, email, country, city, address, password])
        wb.save(EXCEL_FILE)

        flash('Sign up successful! Please log in.', 'success')
        return redirect(url_for('login_page'))

    return render_template('signup.html')

if __name__ == '__main__':
    # Replace with the actual input file path
    grocery_data = pd.read_csv('Datasets/Price_Predictions/grocery_items_bavaria.csv')
    grocery_data.rename(columns={'name': 'Items'}, inplace=True)

    # Delete the CSV file if it exists
    csv_file_path = "Datasets/Output/Price_Prediction/Item_lists.csv"
    if os.path.exists(csv_file_path):
        os.remove(csv_file_path)
        print(f"Existing file {csv_file_path} has been deleted.")

    # Load models from the folder
    models_folder = "Model/Price_Prediction/arima"
    with open(os.path.join(models_folder, "arima_models.pkl"), 'rb') as f:
        loaded_models = pickle.load(f)

    # Predict, save, and plot for specific items
    items_to_predict = [
        "Allgäuer Hof-Milch Butter mild gesäuert 250g",
        "Bio Aubergine 1 Stück",
        "Blumenkohl weiß 1 Stück",
        "Broccoli 500g",
        "Eisbergsalat 1 Stück",
        "Galiamelone 1 Stück",
        "Karotten 1kg",
        "Kartoffeln vorwiegend festkochend 2,5kg",
        "Mango vorgereift 1 Stück",
        "Meggle Feine Butter 250g",
        "Orangen 2kg im Netz",
        "REWE Beste Wahl Feinschmecker Hähnchen 1200g",
        "REWE Bio Zucchini 500g",
        "Rewe Beste Wahl Eier aus Freilandhaltung 10 Stück",
        "Rispentomaten ca. 100g",
        "Spitzkohl ca. 1kg",
        "Tafeltrauben hell kernlos 500g",
        "Zitronen 500g im Netz",
        "Zwiebeln 2kg im Netz",
        "ja! Basmati Reis 1kg",
        "ja! H-Milch 3,5% 1",
        "ja! Sonnenblumenöl 1l"

    ]

    for item_name in items_to_predict:
        predict_and_save(item_name, loaded_models, grocery_data)
           
    initialize_excel()
    app.run(debug=True)
